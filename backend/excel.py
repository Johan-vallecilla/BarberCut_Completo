# ══════════════════════════════════════════════════════
#  excel.py — Registro histórico en Excel
# ══════════════════════════════════════════════════════

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

EXCEL_FILE = "registro_barbercut.xlsx"

VERDE_DARK = "00A87E"
BLANCO     = "FFFFFF"
GRIS       = "F2F2F8"


def inicializar_excel():
    """Crea el archivo Excel con sus hojas si no existe."""
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()

        ws_r = wb.active
        ws_r.title = "Reservas"
        encabezados_r = ["ID", "Nombre", "Teléfono", "Barbero",
                         "Servicio", "Precio (COP)", "Fecha", "Hora", "Estado", "Registrado"]
        ws_r.append(encabezados_r)
        _estilo_encabezado(ws_r, 1, len(encabezados_r))

        ws_b = wb.create_sheet("Barberos")
        encabezados_b = ["ID", "Nombre", "Especialidad", "Teléfono", "Estado", "Registrado"]
        ws_b.append(encabezados_b)
        _estilo_encabezado(ws_b, 1, len(encabezados_b))

        wb.save(EXCEL_FILE)
        print(f"[BarberCut] ✅ Excel '{EXCEL_FILE}' creado.")


def _estilo_encabezado(ws, fila: int, cols: int):
    fill = PatternFill("solid", start_color=VERDE_DARK, end_color=VERDE_DARK)
    font = Font(color=BLANCO, bold=True, name="Arial", size=11)
    alin = Alignment(horizontal="center", vertical="center")
    for col in range(1, cols + 1):
        c = ws.cell(row=fila, column=col)
        c.fill = fill
        c.font = font
        c.alignment = alin
    ws.row_dimensions[fila].height = 22


def registrar_reserva_excel(reserva: dict):
    """Agrega una reserva al Excel como registro histórico."""
    try:
        wb  = openpyxl.load_workbook(EXCEL_FILE)
        ws  = wb["Reservas"]
        color = GRIS if ws.max_row % 2 == 0 else BLANCO
        fill  = PatternFill("solid", start_color=color, end_color=color)
        fila  = [
            reserva["id"],      reserva["nombre"],   reserva["telefono"],
            reserva["barbero"], reserva["servicio"],  reserva["precio"],
            reserva["fecha"],   reserva["hora"],      reserva["estado"],
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ]
        ws.append(fila)
        for col in range(1, len(fila) + 1):
            ws.cell(row=ws.max_row, column=col).fill = fill
        wb.save(EXCEL_FILE)
    except Exception as e:
        print(f"[BarberCut] ⚠️  Excel registro reserva: {e}")


def registrar_barbero_excel(barbero: dict):
    """Agrega un barbero al Excel como registro histórico."""
    try:
        wb  = openpyxl.load_workbook(EXCEL_FILE)
        ws  = wb["Barberos"]
        color = GRIS if ws.max_row % 2 == 0 else BLANCO
        fill  = PatternFill("solid", start_color=color, end_color=color)
        fila  = [
            barbero["id"],      barbero["nombre"],       barbero["especialidad"],
            barbero["telefono"], "activo",
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ]
        ws.append(fila)
        for col in range(1, len(fila) + 1):
            ws.cell(row=ws.max_row, column=col).fill = fill
        wb.save(EXCEL_FILE)
    except Exception as e:
        print(f"[BarberCut] ⚠️  Excel registro barbero: {e}")


def actualizar_estado_excel(reserva_id: int, nuevo_estado: str):
    """Actualiza el estado de una reserva en el Excel."""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb["Reservas"]
        for fila in ws.iter_rows(min_row=2):
            if fila[0].value == reserva_id:
                fila[8].value = nuevo_estado
                if nuevo_estado == "cancelada":
                    for c in fila:
                        c.font = Font(color="CC0000", italic=True, name="Arial", size=10)
                break
        wb.save(EXCEL_FILE)
    except Exception as e:
        print(f"[BarberCut] ⚠️  Excel actualizar estado: {e}")